from typing import Any, List, Optional
from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Body, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, case as sa_case, text
from sqlmodel import Session, select

from app.api.deps import get_current_user
from app.database import get_session
from app.models.client import Client
from app.models.user import User, UserRole
from app.models.vente import Vente
from app.models.objectif import Objectif
from app.models.produit import Produit

router = APIRouter()


# ── WKT → GeoJSON helpers ─────────────────────────────────────────────────────

def _find_paren(s: str, start: int) -> int:
    depth = 0
    for i in range(start, len(s)):
        if s[i] == '(':
            depth += 1
        elif s[i] == ')':
            depth -= 1
            if depth == 0:
                return i
    return -1


def _parse_ring(s: str) -> list:
    s = s.strip().strip('()')
    result = []
    for pair in s.split(','):
        parts = pair.strip().split()
        if len(parts) >= 2:
            result.append([float(parts[0]), float(parts[1])])
    return result


def _parse_polygon(s: str) -> list:
    s = s.strip().strip('()')
    rings, i = [], 0
    while i < len(s):
        if s[i] == '(':
            end = _find_paren(s, i)
            if end == -1:
                break
            rings.append(_parse_ring(s[i:end + 1]))
            i = end + 1
        else:
            i += 1
    return rings


def _wkt_to_coords(wkt: str) -> list:
    s = wkt.strip()
    is_multi = s.upper().startswith('MULTIPOLYGON')
    if is_multi:
        s = s[len('MULTIPOLYGON'):].strip()
    elif s.upper().startswith('POLYGON'):
        s = s[len('POLYGON'):].strip()
    else:
        return []

    s = s.strip().strip('()')
    polygons, i = [], 0
    while i < len(s):
        if s[i] == '(':
            end = _find_paren(s, i)
            if end == -1:
                break
            polygons.append(_parse_polygon(s[i:end + 1]))
            i = end + 1
        else:
            i += 1

    if not is_multi and len(polygons) == 1:
        # POLYGON → wrap as MultiPolygon
        return polygons
    return polygons


# ─────────────────────────────────────────────────────────────────────────────


@router.get("/periodes", response_model=List[str])
def prevendeur_periodes(
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
) -> Any:
    if not current_user.employe_code:
        return []
    rows = session.exec(
        select(Vente.annee_mois)
        .distinct()
        .where(Vente.code_fdv == current_user.employe_code)
        .order_by(Vente.annee_mois.desc())
    ).all()
    return list(rows)


@router.get("/facturation")
def prevendeur_facturation(
    annee_mois: str = Query(...),
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
) -> Any:
    if not current_user.employe_code:
        return {"fdv_nom": "", "periode": annee_mois, "routes": [], "products": [], "products_meta": {}, "total_clients": 0}

    rows = session.exec(
        select(Vente)
        .where(Vente.code_fdv == current_user.employe_code)
        .where(Vente.annee_mois == annee_mois)
        .where(Vente.famille.ilike('sucre') | Vente.famille.ilike('huile'))
    ).all()

    # Exclude explicit BackOffice rows
    rows = [r for r in rows if r.source != 'BackOffice']

    if not rows:
        return {"fdv_nom": current_user.full_name, "periode": annee_mois, "routes": [], "products": [], "products_meta": {}, "total_clients": 0}

    # Product metadata
    codes = {r.code_produit for r in rows if r.code_produit}
    produits_db = session.exec(select(Produit).where(Produit.code_produit.in_(codes))).all()
    code_to_produit = {p.code_produit: p for p in produits_db}
    code_to_label = {p.code_produit: (p.nom_produit or p.description_produit) for p in produits_db}

    # Exclude non-facturable
    rows = [
        r for r in rows
        if not (r.code_produit and r.code_produit in code_to_produit and not code_to_produit[r.code_produit].facturable)
    ]

    def display_label(r: Vente) -> str:
        if r.code_produit and r.code_produit in code_to_label and code_to_label[r.code_produit]:
            return code_to_label[r.code_produit]
        return r.description_produit or ''

    products = sorted({display_label(r) for r in rows if r.description_produit})

    # Build meta map in one pass — O(rows) instead of O(rows × products)
    label_meta: dict = {}
    for r in rows:
        label = display_label(r)
        if label not in label_meta:
            famille = (r.famille or '').lower()
            if r.code_produit and r.code_produit in code_to_produit:
                p = code_to_produit[r.code_produit]
                label_meta[label] = {"uom_vente": p.uom_vente, "colisage": p.colisage, "famille": famille, "prix": p.prix_dd}
            else:
                label_meta[label] = {"uom_vente": None, "colisage": None, "famille": famille, "prix": None}
    products_meta = {p: label_meta.get(p, {"uom_vente": None, "colisage": None, "famille": None, "prix": None}) for p in products}

    # Aggregate: client -> date_label -> product -> qty
    agg: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    client_meta: dict = {}

    for r in rows:
        if not r.nom_client or not r.description_produit or not r.date_commande:
            continue
        label = r.date_commande.strftime('%d/%m/%y')
        agg[r.nom_client][label][display_label(r)] += r.qte_facturee or 0
        if r.nom_client not in client_meta:
            client_meta[r.nom_client] = {
                'code_client': r.code_client,
                'route': r.route or 'Sans route',
            }

    all_dates = sorted({r.date_commande for r in rows if r.date_commande})
    date_labels = [d.strftime('%d/%m/%y') for d in all_dates]

    # Group clients by route
    route_clients: dict = defaultdict(list)
    for nom, meta in sorted(client_meta.items()):
        route_clients[meta['route']].append(nom)

    fdv_nom = rows[0].nom_fdv if rows else current_user.full_name

    # Fetch nom_sodichn from clients table for all client codes
    client_codes = [m['code_client'] for m in client_meta.values() if m.get('code_client')]
    clients_db = session.exec(select(Client).where(Client.customer_no.in_(client_codes))).all() if client_codes else []
    nom_sodichn_map = {c.customer_no: c.nom_sodichn for c in clients_db}

    routes_out = []
    for route in sorted(route_clients.keys()):
        clients_out = []
        for nom in route_clients[route]:
            client_dates = [l for l in date_labels if any(agg[nom][l][p] for p in products)]
            if not client_dates:
                continue
            semaines = {
                l: {p: (agg[nom][l][p] if agg[nom][l][p] else None) for p in products}
                for l in client_dates
            }
            totaux = {p: (sum(agg[nom][l][p] for l in client_dates) or None) for p in products}
            code = client_meta[nom]['code_client']
            clients_out.append({
                "nom_client": nom,
                "code_client": code,
                "nom_sodichn": nom_sodichn_map.get(code),
                "derniere_visite": client_dates[-1],
                "weeks": client_dates,
                "semaines": semaines,
                "totaux": totaux,
            })
        if clients_out:
            routes_out.append({"route": route, "clients": clients_out})

    return {
        "fdv_nom": fdv_nom,
        "periode": annee_mois,
        "products": products,
        "products_meta": products_meta,
        "total_clients": sum(len(r["clients"]) for r in routes_out),
        "routes": routes_out,
    }


@router.get("/admin/stats")
def prevendeur_admin_stats(
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
) -> Any:
    prevendeurs = session.exec(
        select(User)
        .where(User.role == UserRole.PREVENDER)
        .where(User.is_active == True)
        .order_by(User.full_name)
    ).all()

    fdv_codes = [pv.employe_code for pv in prevendeurs if pv.employe_code]
    if not fdv_codes:
        return []

    # 1 query: all distinct (code_fdv, code_client, nom_client) across all prevendeurs
    vente_rows = session.exec(
        select(Vente.code_fdv, Vente.code_client, Vente.nom_client)
        .distinct()
        .where(Vente.code_fdv.in_(fdv_codes), Vente.code_client.isnot(None))
    ).all()

    fdv_clients: dict = defaultdict(list)   # code_fdv -> [(code_client, nom_client)]
    all_client_codes: set = set()
    for code_fdv, code_client, nom_client in vente_rows:
        fdv_clients[code_fdv].append((code_client, nom_client))
        all_client_codes.add(code_client)

    # 1 query: last sale date per prevendeur
    last_sale_rows = session.exec(
        select(Vente.code_fdv, func.max(Vente.date_commande))
        .where(Vente.code_fdv.in_(fdv_codes))
        .group_by(Vente.code_fdv)
    ).all()
    last_sale_map = {code_fdv: last_sale for code_fdv, last_sale in last_sale_rows}

    # 1 query: all client records (for nom_sodichn + updated_at)
    sodichn_map: dict = {}   # customer_no -> {nom_sodichn, updated_at}
    if all_client_codes:
        clients_db = session.exec(
            select(Client).where(Client.customer_no.in_(all_client_codes))
        ).all()
        sodichn_map = {
            c.customer_no: {"nom_sodichn": c.nom_sodichn, "updated_at": c.updated_at}
            for c in clients_db
        }

    today = datetime.now(timezone.utc).date()

    result = []
    for pv in prevendeurs:
        if not pv.employe_code:
            continue
        clients = fdv_clients.get(pv.employe_code, [])
        total_clients = len(clients)

        # Per-client stats
        updated_today = 0
        last_sodichn_date = None
        matched = 0
        clients_detail = []

        for code, nom in sorted(clients, key=lambda x: x[1] or ''):
            info = sodichn_map.get(code, {})
            nom_sodichn = info.get("nom_sodichn")
            updated_at = info.get("updated_at")
            if nom_sodichn:
                matched += 1
            if updated_at and nom_sodichn:
                updated_date = updated_at.date() if hasattr(updated_at, 'date') else updated_at
                if updated_date == today:
                    updated_today += 1
                if last_sodichn_date is None or updated_date > last_sodichn_date:
                    last_sodichn_date = updated_date
            clients_detail.append({
                "code_client": code,
                "nom_client": nom,
                "nom_sodichn": nom_sodichn,
                "updated_at": updated_at.strftime('%Y-%m-%d') if updated_at else None,
            })

        # Count updated on last active day
        updated_on_last_day = 0
        if last_sodichn_date:
            updated_on_last_day = sum(
                1 for c in clients_detail
                if c["updated_at"] and c["updated_at"][:10] == last_sodichn_date.strftime('%Y-%m-%d')
            )

        last_sale = last_sale_map.get(pv.employe_code)
        result.append({
            "id": pv.id,
            "full_name": pv.full_name,
            "employe_code": pv.employe_code,
            "total_clients": total_clients,
            "clients_with_sodichn": matched,
            "remaining": total_clients - matched,
            "completion_pct": round(matched / total_clients * 100) if total_clients > 0 else 0,
            "updated_today": updated_today,
            "last_sodichn_date": last_sodichn_date.strftime('%Y-%m-%d') if last_sodichn_date else None,
            "updated_on_last_day": updated_on_last_day,
            "last_activity": last_sale.strftime('%Y-%m-%d') if last_sale else None,
            "clients": clients_detail,
        })

    return result


@router.get("/admin/stats/export")
def export_clients_excel(
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    prevendeurs = session.exec(
        select(User)
        .where(User.role == UserRole.PREVENDER)
        .where(User.is_active == True)
        .order_by(User.full_name)
    ).all()

    fdv_codes = [pv.employe_code for pv in prevendeurs if pv.employe_code]
    if not fdv_codes:
        return StreamingResponse(BytesIO(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    vente_rows = session.exec(
        select(Vente.code_fdv, Vente.code_client, Vente.nom_client)
        .distinct()
        .where(Vente.code_fdv.in_(fdv_codes), Vente.code_client.isnot(None))
    ).all()

    fdv_clients: dict = defaultdict(list)
    all_client_codes: set = set()
    for code_fdv, code_client, nom_client in vente_rows:
        fdv_clients[code_fdv].append((code_client, nom_client))
        all_client_codes.add(code_client)

    sodichn_map: dict = {}
    if all_client_codes:
        clients_db = session.exec(
            select(Client).where(Client.customer_no.in_(all_client_codes))
        ).all()
        sodichn_map = {
            c.customer_no: {"nom_sodichn": c.nom_sodichn, "updated_at": c.updated_at}
            for c in clients_db
        }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients RC"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="2563EB")
    fdv_font = Font(bold=True, size=10)
    fdv_fill = PatternFill("solid", fgColor="DBEAFE")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["Prévendeur", "Code FDV", "Code Client", "Nom Client", "Nom RC (Sodichn)", "Statut", "Mis à jour le"]
    col_widths = [28, 14, 14, 32, 32, 12, 16]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22

    row_num = 2
    today = datetime.now(timezone.utc).date()

    for pv in prevendeurs:
        if not pv.employe_code:
            continue
        clients = fdv_clients.get(pv.employe_code, [])
        for code, nom in sorted(clients, key=lambda x: x[1] or ''):
            info = sodichn_map.get(code, {})
            nom_sodichn = info.get("nom_sodichn") or ""
            updated_at = info.get("updated_at")
            statut = "✓ Complété" if nom_sodichn else "— Vide"
            updated_str = updated_at.strftime('%Y-%m-%d') if updated_at else ""

            row_data = [pv.full_name, pv.employe_code, code, nom, nom_sodichn, statut, updated_str]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                cell.font = Font(size=9)
                if col == 6:
                    cell.alignment = center
                    if nom_sodichn:
                        cell.font = Font(size=9, color="16A34A", bold=True)
                    else:
                        cell.font = Font(size=9, color="9CA3AF")

            row_num += 1

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"clients_rc_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/admin/drilldown")
def prevendeur_admin_drilldown(
    annee_mois: str = Query(...),
    code_fdv: Optional[str] = Query(None),
    canal: Optional[str] = Query(None),   # "VD" or "VH"
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
) -> Any:
    all_periods = session.exec(
        select(Vente.annee_mois).distinct().order_by(Vente.annee_mois.desc())
    ).all()
    all_periods_list = list(all_periods)

    # Prevendeurs list with their totals for the current period (always unfiltered)
    prevendeurs_db = session.exec(
        select(User)
        .where(User.role == UserRole.PREVENDER)
        .where(User.is_active == True)
        .order_by(User.full_name)
    ).all()
    fdv_name_map = {p.employe_code: p.full_name for p in prevendeurs_db if p.employe_code}

    # Normalize qte_livree: UN rows → divide by colisage; pack rows (CARTON, FARDEAU, …) → keep as-is
    _norm = sa_case(
        (
            (Vente.uom_vente == 'UN') & Produit.colisage.isnot(None),
            Vente.qte_livree / Produit.colisage,
        ),
        else_=Vente.qte_livree,
    )

    fdv_totals_q = (
        select(Vente.code_fdv, func.sum(_norm))
        .outerjoin(Produit, Vente.code_produit == Produit.code_produit)
        .where(Vente.annee_mois == annee_mois)
        .where(Vente.statut_commande == 'Facturé')
        .where(Vente.code_fdv.isnot(None))
        .group_by(Vente.code_fdv)
    )
    if canal:
        fdv_totals_q = fdv_totals_q.where(Vente.canal == canal)
    fdv_totals = {row[0]: round(row[1] or 0) for row in session.exec(fdv_totals_q).all()}

    canal_upper = canal.upper() if canal else None

    # Previous period — use the previous available period in the data (not necessarily month-1)
    try:
        cur_idx = all_periods_list.index(annee_mois)
    except ValueError:
        cur_idx = 0
    prev_periode = all_periods_list[cur_idx + 1] if cur_idx + 1 < len(all_periods_list) else None
    trend_periods = list(reversed(all_periods_list[cur_idx: cur_idx + 6]))  # chronological

    def fetch_rows(periode: str) -> list:
        q = select(
            Vente.date_commande,
            Vente.famille,
            Vente.sous_famille,
            Vente.code_produit,
            Vente.description_produit,
            _norm.label('qty_norm'),
            Vente.code_fdv,
            Vente.nom_fdv,
            Vente.source,
        ).outerjoin(Produit, Vente.code_produit == Produit.code_produit).where(
            Vente.annee_mois == periode, Vente.statut_commande == 'Facturé'
        )
        if code_fdv:
            q = q.where(Vente.code_fdv == code_fdv)
        if canal:
            q = q.where(Vente.canal == canal)
        return list(session.exec(q).all())

    rows = fetch_rows(annee_mois)
    prev_rows = fetch_rows(prev_periode) if prev_periode else []

    period_totals: dict = defaultdict(float)
    if trend_periods:
        q6 = (
            select(Vente.annee_mois, func.sum(_norm))
            .outerjoin(Produit, Vente.code_produit == Produit.code_produit)
            .where(Vente.annee_mois.in_(trend_periods))
            .where(Vente.statut_commande == 'Facturé')
        )
        if code_fdv:
            q6 = q6.where(Vente.code_fdv == code_fdv)
        if canal:
            q6 = q6.where(Vente.canal == canal)
        q6 = q6.group_by(Vente.annee_mois)
        for periode, total in session.exec(q6).all():
            period_totals[periode] = total or 0

    trend_6m = [round(period_totals.get(p, 0)) for p in trend_periods]
    trend_6m_labels = trend_periods

    # Product label resolution
    codes = {r.code_produit for r in rows if r.code_produit}
    produits_db = session.exec(select(Produit).where(Produit.code_produit.in_(codes))).all() if codes else []
    code_to_produit = {p.code_produit: p for p in produits_db}

    def week_idx(d) -> int:
        if d is None:
            return 0
        day = d.day
        if day <= 7: return 0
        if day <= 14: return 1
        if day <= 21: return 2
        return 3

    def product_label(r: Vente) -> str:
        if r.code_produit and r.code_produit in code_to_produit:
            p = code_to_produit[r.code_produit]
            return p.nom_produit or r.description_produit or "Autre"
        return r.description_produit or "Autre"

    # famille -> sous_famille -> produit -> [w0,w1,w2,w3]
    hier: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: [0.0] * 4)))
    # famille -> code_fdv -> total
    fdv_by_famille: dict = defaultdict(lambda: defaultdict(float))
    prev_famille_total: dict = defaultdict(float)
    # famille -> sf -> produit -> code_fdv -> total (for per-product FDV panel)
    fdv_by_sf_prod: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(float))))
    # product label -> code_produit (first seen)
    prod_label_to_code: dict = {}

    for r in rows:
        if not r.date_commande or not r.qty_norm:
            continue
        famille = (r.famille or "").strip().lower()
        sf = (r.sous_famille or "Autres").strip()
        prod = product_label(r)
        w = week_idx(r.date_commande)
        hier[famille][sf][prod][w] += r.qty_norm
        if r.code_produit and prod not in prod_label_to_code:
            prod_label_to_code[prod] = r.code_produit
        if r.code_fdv:
            fdv_by_famille[famille][r.code_fdv] += r.qty_norm
            fdv_by_sf_prod[famille][sf][prod][r.code_fdv] += r.qty_norm
        # Supplement fdv name map from row data
        if r.code_fdv and r.nom_fdv and r.code_fdv not in fdv_name_map:
            fdv_name_map[r.code_fdv] = r.nom_fdv

    for r in prev_rows:
        if not r.qty_norm:
            continue
        famille = (r.famille or "").strip().lower()
        prev_famille_total[famille] += r.qty_norm

    # Objective aggregation for this period
    annee_int, mois_int = int(annee_mois.split('-')[0]), int(annee_mois.split('-')[1])

    # Per-product objectives (total + per-tournée)
    obj_prod_rows = session.exec(
        select(
            Objectif.code_produit,
            Objectif.objectif_packs_vd, Objectif.objectif_packs_vh,
            Objectif.objectif_packs_vd_tournee, Objectif.objectif_packs_vh_tournee,
            Objectif.objectif_tonne_vd, Objectif.objectif_tonne_vh,
        )
        .where(Objectif.mois == mois_int, Objectif.annee == annee_int)
    ).all()
    if canal == 'VD':
        obj_by_prod       = {r.code_produit: r.objectif_packs_vd_tournee for r in obj_prod_rows}
        obj_by_prod_total = {r.code_produit: r.objectif_packs_vd for r in obj_prod_rows}
        obj_tonne_by_prod = {r.code_produit: r.objectif_tonne_vd for r in obj_prod_rows}
    elif canal == 'VH':
        obj_by_prod       = {r.code_produit: r.objectif_packs_vh_tournee for r in obj_prod_rows}
        obj_by_prod_total = {r.code_produit: r.objectif_packs_vh for r in obj_prod_rows}
        obj_tonne_by_prod = {r.code_produit: r.objectif_tonne_vh for r in obj_prod_rows}
    else:
        obj_by_prod       = {r.code_produit: (r.objectif_packs_vd_tournee or 0) + (r.objectif_packs_vh_tournee or 0) for r in obj_prod_rows}
        obj_by_prod_total = {r.code_produit: (r.objectif_packs_vd or 0) + (r.objectif_packs_vh or 0) for r in obj_prod_rows}
        obj_tonne_by_prod = {r.code_produit: (r.objectif_tonne_vd or 0) + (r.objectif_tonne_vh or 0) for r in obj_prod_rows}

    # When a single FDV is selected, display per-route targets instead of global totals
    if code_fdv:
        obj_by_prod_total = obj_by_prod

    # Supplement code_to_produit with objective products that had zero sales this period
    missing_obj_codes = {c for c in obj_by_prod if c and c not in code_to_produit}
    if missing_obj_codes:
        extra = session.exec(select(Produit).where(Produit.code_produit.in_(missing_obj_codes))).all()
        for p in extra:
            code_to_produit[p.code_produit] = p

    # Add zero-sale objective products into hier so they appear in the tree with total=0
    # Skip products that already appear in the hierarchy from ventes — vente famille is authoritative
    codes_in_hier = set(prod_label_to_code.values())
    for code_prod, obj in obj_by_prod.items():
        if not obj or not code_prod or code_prod not in code_to_produit:
            continue
        if code_prod in codes_in_hier:
            continue
        produit = code_to_produit[code_prod]
        prod_name = produit.nom_produit or produit.description_produit or code_prod
        famille = (produit.famille or "").strip().lower()
        sf = (produit.sous_famille or "Autres").strip()
        if prod_name not in hier[famille][sf]:
            hier[famille][sf][prod_name]  # creates [0.0, 0.0, 0.0, 0.0] via defaultdict
            prod_label_to_code[prod_name] = code_prod

    # Per-route total: sum of all tournée targets (derived from obj_by_prod, no extra query)
    objectif_per_route = round(sum(v for v in obj_by_prod.values() if v))

    # Per-fdv per-product sales — queried without code_fdv filter so all pills stay accurate
    _fdv_prod_q = (
        select(Vente.code_fdv, Vente.code_produit, func.sum(_norm))
        .outerjoin(Produit, Vente.code_produit == Produit.code_produit)
        .where(Vente.annee_mois == annee_mois, Vente.statut_commande == 'Facturé', Vente.code_fdv.isnot(None), Vente.code_produit.isnot(None))
    )
    if canal:
        _fdv_prod_q = _fdv_prod_q.where(Vente.canal == canal)
    _fdv_prod_q = _fdv_prod_q.group_by(Vente.code_fdv, Vente.code_produit)

    fdv_prod_totals: dict = defaultdict(lambda: defaultdict(float))
    for _fdv, _prod, _tot in session.exec(_fdv_prod_q).all():
        fdv_prod_totals[_fdv][_prod] = _tot or 0.0

    def compute_achievement_pct(code_fdv: str):
        """Simple average of per-product achievement rates for this prevendeur."""
        rates = []
        prod_sales = fdv_prod_totals.get(code_fdv, {})
        for code_prod, obj in obj_by_prod.items():
            if not obj:
                continue
            sales = prod_sales.get(code_prod, 0.0)
            rates.append(min(sales / obj * 100, 100.0))
        if not rates:
            return None
        return round(sum(rates) / len(rates))

    prevendeurs_out = sorted(
        [
            {
                "code": code,
                "nom": fdv_name_map.get(code, code),
                "total": total,
                "achievement_pct": compute_achievement_pct(code),
            }
            for code, total in fdv_totals.items()
            if not canal_upper or canal_upper in code.upper()
        ],
        key=lambda x: x["nom"],
    )

    # CA (chiffre d'affaires) per famille — SUM(qte_livree * prix_unitaire)
    _ca_q = (
        select(Vente.famille, func.sum(Vente.qte_livree * Vente.prix_unitaire))
        .where(Vente.annee_mois == annee_mois, Vente.statut_commande == 'Facturé')
        .where(Vente.prix_unitaire.isnot(None))
        .group_by(Vente.famille)
    )
    if code_fdv:
        _ca_q = _ca_q.where(Vente.code_fdv == code_fdv)
    if canal:
        _ca_q = _ca_q.where(Vente.canal == canal)
    ca_by_famille = {(r[0] or '').strip().lower(): round(r[1] or 0) for r in session.exec(_ca_q).all()}

    if prev_periode:
        _ca_prev_q = (
            select(Vente.famille, func.sum(Vente.qte_livree * Vente.prix_unitaire))
            .where(Vente.annee_mois == prev_periode, Vente.statut_commande == 'Facturé')
            .where(Vente.prix_unitaire.isnot(None))
            .group_by(Vente.famille)
        )
        if code_fdv:
            _ca_prev_q = _ca_prev_q.where(Vente.code_fdv == code_fdv)
        if canal:
            _ca_prev_q = _ca_prev_q.where(Vente.canal == canal)
        ca_prev_by_famille = {(r[0] or '').strip().lower(): round(r[1] or 0) for r in session.exec(_ca_prev_q).all()}
    else:
        ca_prev_by_famille = {}

    familles_out = []
    for famille, sf_map in sorted(hier.items()):
        f_weeks = [0.0] * 4
        sfs_out = []
        for sf, prod_map in sf_map.items():
            sf_weeks = [0.0] * 4
            prods_out = []
            for prod, wks in prod_map.items():
                for i in range(4):
                    sf_weeks[i] += wks[i]
                prod_top_fdv = sorted(
                    [{"code": c, "nom": fdv_name_map.get(c, c), "total": round(t)}
                     for c, t in fdv_by_sf_prod[famille][sf][prod].items()],
                    key=lambda x: -x["total"]
                )
                prod_code = prod_label_to_code.get(prod)
                prod_obj_t = obj_by_prod.get(prod_code) if prod_code else None
                prod_obj   = obj_by_prod_total.get(prod_code) if prod_code else None
                prod_tonne = obj_tonne_by_prod.get(prod_code) if prod_code else None
                prods_out.append({
                    "nom": prod,
                    "total": round(sum(wks)),
                    "weeks": [round(v) for v in wks],
                    "top_fdv": prod_top_fdv,
                    "objectif_packs": round(prod_obj) if prod_obj else None,
                    "objectif_packs_tournee": round(prod_obj_t) if prod_obj_t else None,
                    "objectif_tonne": round(prod_tonne, 3) if prod_tonne else None,
                })
            for i in range(4):
                f_weeks[i] += sf_weeks[i]
            sf_obj_vals = [
                obj_by_prod_total[prod_label_to_code[p["nom"]]]
                for p in prods_out
                if p["nom"] in prod_label_to_code and prod_label_to_code[p["nom"]] in obj_by_prod_total
                and obj_by_prod_total[prod_label_to_code[p["nom"]]]
            ]
            sf_obj = round(sum(sf_obj_vals)) if sf_obj_vals else None
            sf_tonne_vals = [
                obj_tonne_by_prod[prod_label_to_code[p["nom"]]]
                for p in prods_out
                if p["nom"] in prod_label_to_code and prod_label_to_code[p["nom"]] in obj_tonne_by_prod
                and obj_tonne_by_prod[prod_label_to_code[p["nom"]]]
            ]
            sf_tonne = round(sum(sf_tonne_vals), 3) if sf_tonne_vals else None
            sfs_out.append({
                "nom": sf,
                "total": round(sum(sf_weeks)),
                "weeks": [round(v) for v in sf_weeks],
                "produits": sorted(prods_out, key=lambda x: -x["total"]),
                "objectif_packs": sf_obj,
                "objectif_tonne": sf_tonne,
            })

        f_total = round(sum(f_weeks))
        prev_total = round(prev_famille_total.get(famille, 0))
        delta_pct = round((f_total - prev_total) / prev_total * 100) if prev_total > 0 else None

        top_fdv = sorted(
            [{"code": c, "nom": fdv_name_map.get(c, c), "total": round(t)} for c, t in fdv_by_famille[famille].items()],
            key=lambda x: -x["total"]
        )

        # Derive family objective from SF objectives so hierarchy is consistent
        f_obj_vals = [sf["objectif_packs"] for sf in sfs_out if sf["objectif_packs"]]
        f_obj = round(sum(f_obj_vals)) if f_obj_vals else None
        f_tonne_vals = [sf["objectif_tonne"] for sf in sfs_out if sf["objectif_tonne"]]
        f_tonne = round(sum(f_tonne_vals), 3) if f_tonne_vals else None

        familles_out.append({
            "nom": famille,
            "total": f_total,
            "total_prev": prev_total,
            "delta_pct": delta_pct,
            "weeks": [round(v) for v in f_weeks],
            "sous_familles": sorted(sfs_out, key=lambda x: -x["total"]),
            "top_fdv": top_fdv,
            "objectif_packs": f_obj,
            "objectif_tonne": f_tonne,
            "ca": ca_by_famille.get(famille) or None,
            "ca_prev": ca_prev_by_famille.get(famille) or None,
        })

    # True global objective totals — summed directly from the objectifs table
    if canal == 'VD':
        global_obj_tonne = sum(r.objectif_tonne_vd or 0 for r in obj_prod_rows)
        global_obj_packs = sum(r.objectif_packs_vd or 0 for r in obj_prod_rows)
    elif canal == 'VH':
        global_obj_tonne = sum(r.objectif_tonne_vh or 0 for r in obj_prod_rows)
        global_obj_packs = sum(r.objectif_packs_vh or 0 for r in obj_prod_rows)
    else:
        global_obj_tonne = sum((r.objectif_tonne_vd or 0) + (r.objectif_tonne_vh or 0) for r in obj_prod_rows)
        global_obj_packs = sum((r.objectif_packs_vd or 0) + (r.objectif_packs_vh or 0) for r in obj_prod_rows)

    return {
        "periode": annee_mois,
        "periodes": list(all_periods),
        "prevendeurs": prevendeurs_out,
        "trend_6m": trend_6m,
        "trend_6m_labels": trend_6m_labels,
        "familles": sorted(familles_out, key=lambda x: -x["total"]),
        "objectif_packs_per_route": objectif_per_route or None,
        "global_objectif_tonne": round(global_obj_tonne, 3) if global_obj_tonne else None,
        "global_objectif_packs": round(global_obj_packs) if global_obj_packs else None,
        "global_ca": round(sum(ca_by_famille.values())) if ca_by_famille else None,
    }


@router.get("/admin/analytics")
def prevendeur_admin_analytics(
    annee_mois: str = Query(...),
    famille: Optional[str] = Query(None),
    fdv: Optional[str] = Query(None),
    canal: Optional[str] = Query(None),
    commune: Optional[str] = Query(None),
    unite: str = Query('tonnes'),
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
) -> Any:
    all_periods = session.exec(
        select(Vente.annee_mois).distinct().order_by(Vente.annee_mois.desc())
    ).all()
    all_periods_list = list(all_periods)

    if unite == 'tonnes':
        _norm = sa_case(
            (Produit.poids_unite_vente.isnot(None), Vente.qte_livree * Produit.poids_unite_vente),
            else_=0,
        )
    else:  # packs (default)
        _norm = sa_case(
            (
                (Vente.uom_vente == 'UN') & Produit.colisage.isnot(None),
                Vente.qte_livree / Produit.colisage,
            ),
            else_=Vente.qte_livree,
        )

    def apply_filters(q, include_famille=True, include_commune=True):
        q = q.outerjoin(Produit, Vente.code_produit == Produit.code_produit)
        q = q.where(Vente.annee_mois == annee_mois, Vente.statut_commande == 'Facturé')
        if include_famille and famille:
            q = q.where(Vente.famille.ilike(famille))
        if fdv:
            q = q.where(Vente.code_fdv == fdv)
        if canal:
            q = q.where(Vente.canal == canal)
        if include_commune and commune:
            q = q.where(Vente.commune == commune)
        return q

    # KPIs
    kpi_row = session.exec(apply_filters(
        select(func.coalesce(func.sum(_norm), 0))
    )).one()
    total_ventes = round(kpi_row or 0)

    nb_fdvs = len(session.exec(apply_filters(
        select(Vente.code_fdv).distinct().where(Vente.code_fdv.isnot(None))
    )).all())

    top_famille_row = session.exec(apply_filters(
        select(Vente.famille, func.coalesce(func.sum(_norm), 0))
        .where(Vente.famille.isnot(None))
        .group_by(Vente.famille)
        .order_by(func.sum(_norm).desc())
        .limit(1),
        include_famille=False,
    )).first()

    fdv_name_map = {
        p.employe_code: p.full_name
        for p in session.exec(
            select(User).where(User.role == UserRole.PREVENDER, User.is_active == True)
        ).all()
        if p.employe_code
    }

    top_fdv_row = session.exec(apply_filters(
        select(Vente.code_fdv, func.coalesce(func.sum(_norm), 0))
        .where(Vente.code_fdv.isnot(None))
        .group_by(Vente.code_fdv)
        .order_by(func.sum(_norm).desc())
        .limit(1)
    )).first()

    # Monthly trend (last 6 periods)
    try:
        cur_idx = all_periods_list.index(annee_mois)
    except ValueError:
        cur_idx = 0
    trend_periods = list(reversed(all_periods_list[cur_idx:cur_idx + 6]))

    monthly = []
    if trend_periods:
        def trend_q(q):
            q = q.outerjoin(Produit, Vente.code_produit == Produit.code_produit)
            q = q.where(Vente.annee_mois.in_(trend_periods), Vente.statut_commande == 'Facturé')
            if famille:
                q = q.where(Vente.famille.ilike(famille))
            if fdv:
                q = q.where(Vente.code_fdv == fdv)
            if canal:
                q = q.where(Vente.canal == canal)
            return q

        trend_rows = {
            r[0]: (round(r[1] or 0), r[2] or 0)
            for r in session.exec(trend_q(
                select(
                    Vente.annee_mois,
                    func.coalesce(func.sum(_norm), 0),
                    func.count(Vente.code_fdv),
                ).group_by(Vente.annee_mois)
            )).all()
        }
        monthly = [
            {"month": p, "total": trend_rows.get(p, (0, 0))[0], "nb_fdvs": trend_rows.get(p, (0, 0))[1]}
            for p in trend_periods
        ]

    # By famille (unfiltered by famille so bars always show)
    by_famille = [
        {"famille": r[0], "total": round(r[1] or 0)}
        for r in session.exec(apply_filters(
            select(Vente.famille, func.coalesce(func.sum(_norm), 0))
            .where(Vente.famille.isnot(None))
            .group_by(Vente.famille)
            .order_by(func.sum(_norm).desc()),
            include_famille=False,
        )).all()
    ]

    # Top 10 produits (with famille filter)
    by_produit = [
        {"nom": r[0] or "?", "code": r[1], "total": round(r[2] or 0)}
        for r in session.exec(apply_filters(
            select(Vente.description_produit, Vente.code_produit, func.coalesce(func.sum(_norm), 0))
            .where(Vente.description_produit.isnot(None))
            .group_by(Vente.description_produit, Vente.code_produit)
            .order_by(func.sum(_norm).desc())
            .limit(10)
        )).all()
    ]

    # Top 10 FDVs (with famille filter)
    by_fdv = [
        {"nom": fdv_name_map.get(r[0], r[0]), "code": r[0], "total": round(r[1] or 0)}
        for r in session.exec(apply_filters(
            select(Vente.code_fdv, func.coalesce(func.sum(_norm), 0))
            .where(Vente.code_fdv.isnot(None))
            .group_by(Vente.code_fdv)
            .order_by(func.sum(_norm).desc())
            .limit(10)
        )).all()
    ]

    # By location — join ventes with location_communes by commune name
    loc_params: dict = {"loc_annee_mois": annee_mois}
    loc_conditions = [
        "v.annee_mois = :loc_annee_mois",
        "v.statut_commande = 'Facturé'",
        "v.commune IS NOT NULL",
    ]
    if famille:
        loc_conditions.append("LOWER(v.famille) = LOWER(:loc_famille)")
        loc_params["loc_famille"] = famille
    if fdv:
        loc_conditions.append("v.code_fdv = :loc_fdv")
        loc_params["loc_fdv"] = fdv
    if canal:
        loc_conditions.append("v.canal = :loc_canal")
        loc_params["loc_canal"] = canal

    if unite == 'tonnes':
        loc_norm = "CASE WHEN p.poids_unite_vente IS NOT NULL THEN v.qte_livree * p.poids_unite_vente ELSE 0 END"
    else:
        loc_norm = "CASE WHEN v.uom_vente = 'UN' AND p.colisage IS NOT NULL THEN v.qte_livree / p.colisage ELSE v.qte_livree END"
    where_clause = " AND ".join(loc_conditions)
    loc_sql = f"""
        SELECT lc.commune_code, lc.commune_name,
               COALESCE(SUM({loc_norm}), 0) AS total
        FROM ventes v
        LEFT JOIN produits p ON v.code_produit = p.code_produit
        JOIN location_communes lc ON LOWER(v.commune) = LOWER(lc.commune_name)
        WHERE {where_clause}
        GROUP BY lc.commune_code, lc.commune_name
        ORDER BY total DESC
    """
    loc_rows = session.execute(text(loc_sql), loc_params).all()
    by_location = [
        {"code": row[0], "name": row[1], "total": round(row[2] or 0)}
        for row in loc_rows
    ]

    return {
        "kpis": {
            "total_ventes": total_ventes,
            "nb_fdvs": nb_fdvs,
            "top_famille": {"nom": top_famille_row[0], "total": round(top_famille_row[1])}
                           if top_famille_row else None,
            "top_fdv": {
                "nom": fdv_name_map.get(top_fdv_row[0], top_fdv_row[0]),
                "code": top_fdv_row[0],
                "total": round(top_fdv_row[1]),
            } if top_fdv_row else None,
        },
        "monthly": monthly,
        "by_famille": by_famille,
        "by_produit": by_produit,
        "by_fdv": by_fdv,
        "by_location": by_location,
        "periodes": all_periods_list,
    }


@router.get("/admin/communes-geojson")
def admin_communes_geojson(
    codes: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
) -> Any:
    import json as _json
    sql = "SELECT commune_code, commune_name, ST_AsGeoJSON(geom) FROM location_communes WHERE geom IS NOT NULL"
    params: dict = {}
    if codes:
        code_list = [int(c) for c in codes.split(",") if c.strip().isdigit()]
        if code_list:
            sql += " AND commune_code = ANY(:codes)"
            params["codes"] = code_list
    sql += " ORDER BY commune_code"
    rows = session.execute(text(sql), params).all()
    features = [
        {"type": "Feature", "properties": {"code": code, "name": name}, "geometry": _json.loads(geom_json)}
        for code, name, geom_json in rows if geom_json
    ]
    return {"type": "FeatureCollection", "features": features}


@router.get("/objectifs")
def prevendeur_objectifs(
    annee_mois: str = Query(...),
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
) -> Any:
    if not current_user.employe_code:
        return []

    code_fdv = current_user.employe_code
    annee_int, mois_int = int(annee_mois.split('-')[0]), int(annee_mois.split('-')[1])

    code_upper = code_fdv.upper()
    use_vd = 'VH' not in code_upper  # default to VD unless explicitly VH

    # Actual sales per product — with name fallback from Vente
    sales_rows = session.exec(
        select(Vente.code_produit, Vente.description_produit, func.sum(Vente.qte_facturee))
        .where(Vente.code_fdv == code_fdv)
        .where(Vente.annee_mois == annee_mois)
        .where(Vente.source != 'BackOffice')
        .group_by(Vente.code_produit, Vente.description_produit)
    ).all()

    sales_by_code: dict = {}
    desc_by_code: dict = {}
    for code, desc, qty in sales_rows:
        if not code:
            continue
        sales_by_code[code] = sales_by_code.get(code, 0) + (qty or 0)
        if desc and code not in desc_by_code:
            desc_by_code[code] = desc

    # Objectives for this period
    obj_rows = session.exec(
        select(Objectif)
        .where(Objectif.mois == mois_int, Objectif.annee == annee_int)
    ).all()

    obj_by_code: dict = {}
    for obj in obj_rows:
        if not obj.code_produit:
            continue
        objectif = obj.objectif_packs_vd_tournee if use_vd else obj.objectif_packs_vh_tournee
        obj_by_code[obj.code_produit] = objectif or 0

    all_codes = set(sales_by_code.keys()) | set(obj_by_code.keys())
    produits_db = session.exec(select(Produit).where(Produit.code_produit.in_(all_codes))).all() if all_codes else []
    code_to_produit = {p.code_produit: p for p in produits_db}

    result = []
    for code in all_codes:
        actual = sales_by_code.get(code, 0)
        objectif = obj_by_code.get(code, 0)
        produit = code_to_produit.get(code)
        nom = (produit.nom_produit or produit.description_produit if produit else None) or desc_by_code.get(code) or code
        famille = (produit.famille or '').lower() if produit else ''
        pct = round(min(actual / objectif * 100, 100)) if objectif else 0
        result.append({
            "code_produit": code,
            "nom_produit": nom,
            "famille": famille,
            "actual": round(actual),
            "objectif": round(objectif),
            "pct": pct,
        })

    return sorted(result, key=lambda x: (x['pct'], x['nom_produit']))


@router.patch("/clients/{code_client}")
def update_nom_sodichn(
    code_client: str,
    nom_sodichn: str = Body(..., embed=True),
    nom_client: str = Body(..., embed=True),
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
) -> Any:
    client = session.exec(select(Client).where(Client.customer_no == code_client)).first()
    if client:
        client.nom_sodichn = nom_sodichn or None
        client.updated_at = datetime.now(timezone.utc)
    else:
        client = Client(customer_no=code_client, name=nom_client, nom_sodichn=nom_sodichn or None)
        session.add(client)
    session.commit()
    return {"ok": True}
