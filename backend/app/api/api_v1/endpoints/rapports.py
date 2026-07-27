from typing import Any, List, Optional
from collections import defaultdict
from datetime import date, timedelta
import calendar
import io
import re
import zipfile

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select

from app.api.deps import get_current_user
from app.database import get_session
from app.models.vente import Vente
from app.models.produit import Produit

router = APIRouter()

FAMILLES_RAPPORT = ['sucre', 'huile']




@router.get("/facturation-clients", response_model=List[str])
def list_facturation_clients(
    annee_mois: str = Query(...),
    nom_fdv: str = Query(...),
    session: Session = Depends(get_session),
) -> Any:
    """Return distinct clients for a given FDV/period that have sucre or huile sales."""
    rows = session.exec(
        select(Vente.nom_client)
        .distinct()
        .where(Vente.annee_mois == annee_mois)
        .where(Vente.nom_fdv == nom_fdv)
        .where(Vente.famille.ilike('sucre') | Vente.famille.ilike('huile'))
        .order_by(Vente.nom_client)
    ).all()
    return [r for r in rows if r]


@router.get("/facturation")
def get_rapport_facturation(
    annee_mois: str = Query(...),
    nom_fdv: str = Query(...),
    clients: Optional[List[str]] = Query(default=None),
    session: Session = Depends(get_session),
    current_user: Any = Depends(get_current_user),
) -> Any:
    year, month = int(annee_mois[:4]), int(annee_mois[5:7])

    rows = session.exec(
        select(Vente)
        .where(Vente.annee_mois == annee_mois)
        .where(Vente.nom_fdv == nom_fdv)
        .where(Vente.famille.ilike('sucre') | Vente.famille.ilike('huile'))
    ).all()

    if clients:
        rows = [r for r in rows if r.nom_client in clients]

    # Build code_produit -> display name map (nom_produit if set, else description_produit)
    codes = {r.code_produit for r in rows if r.code_produit}
    produits = session.exec(select(Produit).where(Produit.code_produit.in_(codes))).all()
    code_to_label = {p.code_produit: (p.nom_produit or p.description_produit) for p in produits}
    code_to_produit = {p.code_produit: p for p in produits}

    def display_label(r: Vente) -> str:
        if r.code_produit and r.code_produit in code_to_label and code_to_label[r.code_produit]:
            return code_to_label[r.code_produit]
        return r.description_produit or ''

    # Distinct products (columns), sorted
    products = sorted({display_label(r) for r in rows if r.description_produit})

    def meta_for_label(label: str) -> dict:
        for r in rows:
            if display_label(r) == label:
                famille = (r.famille or '').lower()
                if r.code_produit and r.code_produit in code_to_produit:
                    p = code_to_produit[r.code_produit]
                    return {"uom_vente": p.uom_vente, "colisage": p.colisage, "famille": famille}
                return {"uom_vente": None, "colisage": None, "famille": famille}
        return {"uom_vente": None, "colisage": None, "famille": None}

    products_meta = {p: meta_for_label(p) for p in products}

    # Distinct clients (rows), sorted
    client_names = sorted({r.nom_client for r in rows if r.nom_client})

    # Distinct dates that have actual orders, sorted
    dates = sorted({r.date_commande for r in rows if r.date_commande})
    date_labels = [d.strftime('%d/%m/%y') for d in dates]

    # Aggregate: client -> date_label -> product -> qty
    agg: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in rows:
        if not r.nom_client or not r.description_produit or not r.date_commande:
            continue
        label = r.date_commande.strftime('%d/%m/%y')
        agg[r.nom_client][label][display_label(r)] += r.qte_facturee or 0

    clients_out = []
    for nom in client_names:
        # Only keep dates where this client has at least one product qty
        client_dates = [
            label for label in date_labels
            if any(agg[nom][label][p] for p in products)
        ]
        semaines = {}
        for label in client_dates:
            semaines[label] = {
                p: (agg[nom][label][p] if agg[nom][label][p] else None)
                for p in products
            }
        totaux = {
            p: (sum(agg[nom][label][p] for label in client_dates) or None)
            for p in products
        }
        clients_out.append({
            "nom_client": nom,
            "semaines": semaines,
            "totaux": totaux,
            "weeks": client_dates,
        })

    return {
        "fdv": nom_fdv,
        "periode": annee_mois,
        "weeks": date_labels,
        "products": products,
        "products_meta": products_meta,
        "clients": clients_out,
    }


@router.get("/export-clients-zip")
def export_clients_zip(
    annee_mois: str = Query(...),
    nom_fdv: str = Query(...),
    clients: Optional[List[str]] = Query(default=None),
    display_mode: str = Query(default="brut"),
    session: Session = Depends(get_session),
    current_user: Any = Depends(get_current_user),
) -> StreamingResponse:
    """One .xlsx per client (all products), bundled as ZIP."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    rows = session.exec(
        select(Vente)
        .where(Vente.annee_mois == annee_mois)
        .where(Vente.nom_fdv == nom_fdv)
        .where(Vente.famille.ilike('sucre') | Vente.famille.ilike('huile'))
    ).all()

    if clients:
        rows = [r for r in rows if r.nom_client in clients]

    # Build code_produit → Produit map
    codes = {r.code_produit for r in rows if r.code_produit}
    produits_db = session.exec(select(Produit).where(Produit.code_produit.in_(codes))).all()
    code_to_produit = {p.code_produit: p for p in produits_db}

    def product_label(r: Vente) -> str:
        p = code_to_produit.get(r.code_produit or "")
        if p and (p.nom_produit or p.description_produit):
            return p.nom_produit or p.description_produit
        return r.description_produit or r.code_produit or ""

    def colisage_for(code: str) -> Optional[float]:
        p = code_to_produit.get(code or "")
        return p.colisage if p else None

    use_unites = display_mode == "unites"

    # Aggregate: client → product_key → {label, qty, prix}
    agg: dict = defaultdict(lambda: defaultdict(
        lambda: {"label": "", "qty": 0.0, "prix": None, "colisage": None}
    ))
    for r in rows:
        if not r.nom_client:
            continue
        key = r.code_produit or r.description_produit or ""
        e = agg[r.nom_client][key]
        e["label"] = product_label(r)
        e["qty"] += r.qte_facturee or 0
        if e["prix"] is None:
            prod = code_to_produit.get(r.code_produit or "")
            if prod and prod.prix:
                e["prix"] = prod.prix
        if e["colisage"] is None:
            e["colisage"] = colisage_for(r.code_produit or "")

    def safe_name(s: str) -> str:
        return re.sub(r'[\\/*?:"<>|]', "_", s)[:80]

    # Styles
    HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    TTL_FILL = PatternFill("solid", fgColor="D6E4F0")
    TTL_FONT = Font(bold=True, size=11)
    THIN = Side(border_style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center")
    MID = Alignment(vertical="center")

    period_label = f"{annee_mois[5:7]}/{annee_mois[:4]}"

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for client_name, products in agg.items():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Détail"

            # Title rows
            ws.append([client_name])
            ws.merge_cells("A1:D1")
            ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
            ws["A1"].alignment = CENTER
            ws.row_dimensions[1].height = 22

            ws.append([f"Prévendeur : {nom_fdv}   |   Période : {period_label}"])
            ws.merge_cells("A2:D2")
            ws["A2"].font = Font(italic=True, size=10, color="555555")
            ws["A2"].alignment = CENTER
            ws.row_dimensions[2].height = 15

            ws.append([])  # blank spacer

            # Header row (row 4)
            qty_header = "Quantité (unités)" if use_unites else "Quantité (colis)"
            headers = ["Code Produit", qty_header, "Prix Unitaire", "Total"]
            ws.append(headers)
            for cell in ws[4]:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = CENTER
                cell.border = BORDER
            ws.row_dimensions[4].height = 18

            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 16

            # Data rows
            grand_qty = 0.0
            grand_total = 0.0
            for key in sorted(products.keys(), key=lambda k: products[k]["label"]):
                e = products[key]
                if e["qty"] == 0:
                    continue
                prix = e["prix"]
                qty = e["qty"]
                if use_unites and e["colisage"]:
                    qty = qty * e["colisage"]
                qty = round(qty, 2)
                line_total = round(qty * prix, 2) if prix else None
                ws.append([e["label"], qty, prix, line_total])
                row = ws[ws.max_row]
                for cell in row:
                    cell.border = BORDER
                    cell.alignment = MID
                row[1].alignment = CENTER  # qty centered
                if prix is not None:
                    row[2].number_format = "#,##0.00"
                if line_total is not None:
                    row[3].number_format = "#,##0.00"
                grand_qty += e["qty"]
                grand_total += line_total or 0

            # Total row
            ws.append(["TOTAL", grand_qty, "", round(grand_total, 2)])
            total_row = ws[ws.max_row]
            for cell in total_row:
                cell.fill = TTL_FILL
                cell.font = TTL_FONT
                cell.border = BORDER
                cell.alignment = MID
            total_row[1].alignment = CENTER
            total_row[3].number_format = "#,##0.00"

            xl_buf = io.BytesIO()
            wb.save(xl_buf)
            xl_buf.seek(0)
            zf.writestr(f"{safe_name(client_name)}.xlsx", xl_buf.read())

    zip_buf.seek(0)
    zip_name = f"export_{safe_name(nom_fdv)}_{annee_mois}.zip"
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{zip_name}"'},
    )
