from typing import Any, List, Optional
from datetime import datetime, timezone
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select

from app.api.deps import get_current_user
from app.database import get_session
from app.models.client import Client, ClientCreate, ClientUpdate, ClientRead
from app.models.user import User, UserRole

router = APIRouter()


@router.get("", response_model=List[ClientRead])
def list_clients(
    search: Optional[str] = Query(default=None),
    is_active: Optional[bool] = None,
    session: Session = Depends(get_session),
) -> Any:
    q = select(Client)
    if is_active is not None:
        q = q.where(Client.is_active == is_active)
    if search:
        term = f"%{search}%"
        q = q.where(
            Client.name.ilike(term)
            | Client.phone.ilike(term)
        )
    return session.exec(q.order_by(Client.name)).all()


@router.get("/export")
def export_clients_excel(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    clients = session.exec(select(Client).where(Client.is_active == True).order_by(Client.name)).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", vertical="center")

    headers = ["N° Client", "Nom", "Nom RC (Sodichn)", "Wilaya", "Commune", "Téléphone"]
    widths  = [18,          30,    32,                  18,       18,        16]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    cat_labels = {"gros": "Gros", "detail": "Détail", "horeca": "Horeca"}

    for row_i, c in enumerate(clients, 2):
        row_data = [
            c.customer_no or "", c.name, c.nom_sodichn or "", c.wilaya or "", c.commune or "", c.phone or "",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col, value=value)
            cell.border = border
            cell.font = Font(size=9)

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"clients_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("", response_model=ClientRead, status_code=201)
def create_client(
    client_in: ClientCreate,
    session: Session = Depends(get_session),
) -> Any:
    client = Client(**client_in.model_dump())
    session.add(client)
    session.commit()
    session.refresh(client)
    return client


@router.patch("/{client_id}", response_model=ClientRead)
def update_client(
    client_id: int,
    client_in: ClientUpdate,
    session: Session = Depends(get_session),
) -> Any:
    client = session.get(Client, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client introuvable")
    for k, v in client_in.model_dump(exclude_unset=True).items():
        setattr(client, k, v)
    client.updated_at = datetime.now(timezone.utc)
    session.add(client)
    session.commit()
    session.refresh(client)
    return client


@router.delete("/{client_id}", status_code=204)
def delete_client(
    client_id: int,
    session: Session = Depends(get_session),
) -> None:
    client = session.get(Client, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client introuvable")
    session.delete(client)
    session.commit()
